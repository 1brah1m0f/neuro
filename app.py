import io
import re
import os
import unicodedata
from io import BytesIO
from typing import Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Excel Tools", layout="centered")

# ==========================================
# CONSTANTS & HELPER FUNCTIONS
# ==========================================
TEMPLATE_FILE = "report_template.xlsx"

SENTIMENT_MAP = {
    "müsbət": "Positive", "musbet": "Positive", "müsbет": "Positive", "MÜSBƏT": "Positive",
    "pozitiv": "Positive", "POZİTİV": "Positive", "POZITIV": "Positive",
    "pozitif": "Positive", "POZİTİF": "Positive",
    "positive": "Positive", "POSITIVE": "Positive", "Positive": "Positive",
    "müspet": "Positive", "muspet": "Positive",

    "mənfi": "Negative", "menfi": "Negative", "MƏNFİ": "Negative", "MENFİ": "Negative", "MENFI": "Negative",
    "negativ": "Negative", "NEGATİV": "Negative", "NEGATIV": "Negative",
    "neqativ": "Negative", "NEQATİV": "Negative", "NEQATIV": "Negative",
    "negative": "Negative", "NEGATIVE": "Negative", "Negative": "Negative",

    "neytral": "Neutral", "NEYTRAL": "Neutral", "Neytral": "Neutral",
    "neutral": "Neutral", "NEUTRAL": "Neutral", "Neutral": "Neutral",
    "tərəfsiz": "Neutral", "TƏRƏFSİZ": "Neutral", "terefsiz": "Neutral", "TEREFSIZ": "Neutral",

    "1": "Positive", "0": "Neutral", "-1": "Negative",
}

def _fold(s: str) -> str:
    s = unicodedata.normalize("NFKD", s.lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.replace("ı", "i").replace("ə", "e")

def normalize_date_text(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip()
    s = re.sub(r"\s+\d{1,2}:\d{2}(:\d{2})?\s*(AM|PM)?$", "", s, flags=re.I)
    s = s.replace("/", "-").replace(".", "-").replace("_", "-").replace("–", "-").replace("—", "-")
    s = re.sub(r"-{2,}", "-", s)
    return s.strip('"').strip("'").strip()

def parse_dates_robust(series: pd.Series) -> pd.Series:
    raw = series.fillna("").astype(str).str.strip()
    is_numeric = raw.str.isnumeric() & (raw.str.len() >= 4)

    parsed_dayfirst = pd.to_datetime(raw, errors="coerce", dayfirst=True)
    parsed_monthfirst = pd.to_datetime(raw, errors="coerce", dayfirst=False)

    best_parsed = parsed_dayfirst if parsed_dayfirst.notna().sum() >= parsed_monthfirst.notna().sum() else parsed_monthfirst

    if best_parsed.isna().mean() > 0.4:
        norm = raw.map(normalize_date_text)
        p3 = pd.to_datetime(norm, errors="coerce", dayfirst=True)
        p4 = pd.to_datetime(norm, errors="coerce", dayfirst=False)
        best_norm = p3 if p3.notna().sum() >= p4.notna().sum() else p4

        if best_norm.notna().sum() > best_parsed.notna().sum():
            best_parsed = best_norm

    if is_numeric.any():
        numeric_dates = pd.to_datetime(raw[is_numeric].astype(float), unit="D", origin="1899-12-30", errors="coerce")
        best_parsed.loc[is_numeric] = numeric_dates

    return best_parsed.dt.normalize()

def translate_sentiment(x) -> str:
    if pd.isna(x):
        return ""
    raw = str(x).strip()
    if not raw:
        return ""
    key = re.sub(r"\s+", " ", _fold(raw))
    return SENTIMENT_MAP.get(key, raw)

def best_col(df: pd.DataFrame, candidates) -> Optional[str]:
    cols_lower = {str(c).lower(): c for c in df.columns}
    for cand in candidates:
        for k, original in cols_lower.items():
            if cand in k:
                return original
    return None

def _guess_date_col(df: pd.DataFrame, exclude_cols=None) -> Optional[str]:
    exclude = set(str(c).lower() for c in (exclude_cols or []))
    best, best_score = None, 0
    for col in df.columns:
        if str(col).lower() in exclude:
            continue
        try:
            vals = df[col].dropna().astype(str).str.strip()
            if vals.empty:
                continue
            sample = vals.head(50)
            score = parse_dates_robust(sample).notna().sum()
            if score > best_score:
                best_score = score
                best = col
        except Exception:
            continue
    if best and best_score >= max(2, len(df.head(50)) * 0.3):
        return best
    return None

def process_sheet(df: pd.DataFrame) -> pd.DataFrame:
    df = df.reset_index(drop=True)
    url_col = best_col(df, ["url", "link", "href", "source"])
    content_col = best_col(df, ["content", "text", "metn", "mətn", "kontent", "message", "post", "caption", "description", "body"])

    date_col = best_col(df, ["date", "day", "tarix", "data", "datetime", "time", "timestamp", "created", "published", "posted", "vaxt", "zaman", "created_at", "publish", "gun", "gün"])

    if not date_col:
        date_col = _guess_date_col(df, exclude_cols=[url_col or "", content_col or ""])
    sentiment_col = best_col(df, ["sentiment", "hiss", "emosiya", "rating", "tone", "mood", "label", "class"])
    measures_col = best_col(df, ["measures", "tədbir", "action"])

    if not url_col and not content_col:
        raise ValueError(f"URL və Content sütunları tapılmadı. Mövcud sütunlar: {list(df.columns)}")

    if not url_col:
        url_col = content_col
    if not content_col:
        content_col = url_col

    if date_col:
        parsed_dates = parse_dates_robust(df[date_col])
    else:
        parsed_dates = pd.Series(pd.NaT, index=df.index)

    sentiments = df[sentiment_col].map(translate_sentiment).values if sentiment_col else [""] * len(df)

    out = pd.DataFrame({
        "URL": df[url_col].fillna("").astype(str).str.strip(),
        "Content": df[content_col].fillna("").astype(str).str.strip(),
        "Date": parsed_dates,
        "Sentiment": sentiments,
        "_sort": parsed_dates.values,
    })

    if measures_col:
        out["Measures taken"] = df[measures_col].fillna("").astype(str).str.strip()

    out = out.sort_values("_sort", ascending=True, na_position="last").reset_index(drop=True).drop(columns=["_sort"])
    return out

def process_excel(uploaded_bytes: bytes) -> tuple:
    all_sheets = pd.read_excel(io.BytesIO(uploaded_bytes), sheet_name=None, dtype=str)
    skipped, processed = [], {}

    for sheet_name, df in all_sheets.items():
        if sheet_name.lower() == "report":
            continue
        try:
            processed[sheet_name] = process_sheet(df)
        except ValueError as e:
            skipped.append(f"⚠️ Sheet '{sheet_name}' skip olundu: {e}")

    if not processed:
        raise ValueError("Heç bir sheet emal oluna bilmədi.\n" + "\n".join(skipped))

    buf = io.BytesIO(uploaded_bytes)

    with pd.ExcelWriter(buf, engine="openpyxl", mode="a", if_sheet_exists="replace", datetime_format="m/d/yyyy") as writer:
        for sheet_name, cleaned in processed.items():
            cleaned.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            header = {cell.value: cell.column for cell in ws[1]}
            url_col_idx = header.get("URL")
            date_col_idx = header.get("Date")
            hyperlink_font = Font(color="0563C1", underline="single")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if url_col_idx:
                    cell = row[url_col_idx - 1]
                    url_val = str(cell.value or "").strip()
                    if url_val.startswith("http"):
                        cell.hyperlink = url_val
                        cell.font = hyperlink_font
                if date_col_idx:
                    dcell = row[date_col_idx - 1]
                    if dcell.value is not None:
                        dcell.number_format = "m/d/yyyy"

    return buf.getvalue(), skipped

# ==========================================
# TEMPLATE REPORT BUILDER
# ==========================================
def make_unique_sheet_name(workbook, sheet_name: str) -> str:
    existing_names = set(workbook.sheetnames)
    if sheet_name not in existing_names:
        return sheet_name

    base = sheet_name[:25]
    i = 1
    while True:
        candidate = f"{base}_{i}"
        if candidate not in existing_names:
            return candidate
        i += 1

def append_sheets_to_template(data_bytes: bytes) -> bytes:
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template faylı tapılmadı: {TEMPLATE_FILE}")

    template_wb = load_workbook(TEMPLATE_FILE)
    data_sheets = pd.read_excel(BytesIO(data_bytes), sheet_name=None, dtype=str)

    insert_position = 1  # 1-ci sheetdən sonra əlavə et

    for sheet_name, df in data_sheets.items():
        # Əgər data faylında "Report" varsa skip et
        if str(sheet_name).strip().lower() == "report":
            continue

        final_sheet_name = make_unique_sheet_name(template_wb, sheet_name)
        ws = template_wb.create_sheet(title=final_sheet_name, index=insert_position)

        for row in dataframe_to_rows(df.fillna(""), index=False, header=True):
            ws.append(row)

        insert_position += 1

    output = BytesIO()
    template_wb.save(output)
    output.seek(0)
    return output.getvalue()

# ==========================================
# STREAMLIT UI & ROUTING
# ==========================================
tool = st.sidebar.radio(
    "Alət seç:",
    ["Excel Sheet Combiner", "Excel Cleaner", "Template Report Builder"]
)
st.title(tool)

if tool == "Excel Sheet Combiner":
    uploaded_files = st.file_uploader("Upload Excel files", type="xlsx", accept_multiple_files=True)
    sheet_names_input = st.text_input(
        "Enter comma-separated sheet names to combine",
        "Tiktok,Facebook,News,YouTube,Linkedin,Twitter,Instagram"
    )

    if st.button("Combine Files", type="primary"):
        if not uploaded_files:
            st.warning("Please upload at least one Excel file.")
        else:
            sheet_names = [s.strip().lower() for s in sheet_names_input.split(",")]
            combined_data = {sheet: [] for sheet in sheet_names}

            for uploaded_file in uploaded_files:
                try:
                    company_name = os.path.splitext(uploaded_file.name)[0]
                    xls = pd.ExcelFile(uploaded_file)
                    excel_sheets = {s.lower(): s for s in xls.sheet_names}

                    for sheet in sheet_names:
                        if sheet in excel_sheets:
                            df = pd.read_excel(xls, sheet_name=excel_sheets[sheet])
                            df["Company"] = company_name
                            combined_data[sheet].append(df)
                        else:
                            st.warning(f"Sheet '{sheet}' not found in {uploaded_file.name}. Skipping.")
                except Exception as e:
                    st.error(f"Error processing file {uploaded_file.name}: {e}")

            has_data = any(combined_data[s] for s in sheet_names)
            if not has_data:
                st.error("Heç bir sheet-də data tapılmadı.")
            else:
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    for sheet, data in combined_data.items():
                        if data:
                            combined_df = pd.concat(data, ignore_index=True)
                            combined_df.to_excel(writer, sheet_name=sheet.capitalize(), index=False)
                            st.success(f"Sheet '{sheet}' combined successfully.")
                        else:
                            st.warning(f"No data for sheet '{sheet}'.")

                output.seek(0)
                st.download_button(
                    label="Download Combined Excel File",
                    data=output,
                    file_name="combined_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

elif tool == "Excel Cleaner":
    st.write("Excel yüklə → bütün sheet-lər ayrı-ayrı emal olunacaq. Hər sheet-də çıxış: URL, Content, Date (M/D/YYYY), Sentiment.")
    uploaded = st.file_uploader("Excel faylını seç (.xlsx)", type=["xlsx"])

    col1, col2 = st.columns(2)
    with col1:
        run_btn = st.button("Təmizlə və hazırla", type="primary")
    with col2:
        st.caption("Çıxış: hər sheet ayrıca → URL · Content · Date · Sentiment")

    if run_btn:
        if not uploaded:
            st.error("Əvvəl Excel faylını yüklə.")
        else:
            try:
                with st.spinner("Emal olunur..."):
                    result, skipped = process_excel(uploaded.getvalue())
                for msg in skipped:
                    st.warning(msg)
                st.success("Hazırdır ✅")
                st.download_button(
                    label="Cleaned Excel-i yüklə",
                    data=result,
                    file_name="cleaned_final.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"Xəta baş verdi: {str(e)}")

else:
    st.write("Template daxildə sabitdir. Sadəcə data Excel faylını yüklə. Bütün sheet-lər 1-ci sheet-dən sonra əlavə olunacaq.")

    if os.path.exists(TEMPLATE_FILE):
        st.success(f"Template tapıldı: {TEMPLATE_FILE}")
    else:
        st.error(f"Template tapılmadı: {TEMPLATE_FILE}. Faylı app.py ilə eyni qovluğa qoy.")

    data_file = st.file_uploader("Data Excel faylını seç (.xlsx)", type=["xlsx"], key="data_file")

    if st.button("Template ilə birləşdir", type="primary"):
        if not os.path.exists(TEMPLATE_FILE):
            st.error(f"Template faylı yoxdur: {TEMPLATE_FILE}")
        elif not data_file:
            st.error("Əvvəl data faylını yüklə.")
        else:
            try:
                with st.spinner("Template üzərinə sheet-lər əlavə olunur..."):
                    final_file = append_sheets_to_template(data_file.getvalue())

                st.success("Hazırdır ✅")
                st.download_button(
                    label="Final Excel-i yüklə",
                    data=final_file,
                    file_name="template_with_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"Xəta baş verdi: {str(e)}")