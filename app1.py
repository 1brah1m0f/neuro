import io
import re
import unicodedata
from typing import Optional
import pandas as pd
import streamlit as st
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Excel Cleaner", layout="centered")

SENTIMENT_MAP = {
    # Positive variants (lower, upper, mixed)
    "müsbət": "Positive", "musbet": "Positive",
    "müsbет": "Positive", "MÜSBƏT": "Positive",
    "pozitiv": "Positive", "POZİTİV": "Positive", "POZITIV": "Positive",
    "pozitif": "Positive", "POZİTİF": "Positive",
    "positive": "Positive", "POSITIVE": "Positive", "Positive": "Positive",
    "müspet": "Positive", "muspet": "Positive",

    # Negative variants
    "mənfi": "Negative", "menfi": "Negative",
    "MƏNFİ": "Negative", "MENFİ": "Negative", "MENFI": "Negative",
    "negativ": "Negative", "NEGATİV": "Negative", "NEGATIV": "Negative",
    "neqativ": "Negative", "NEQATİV": "Negative", "NEQATIV": "Negative",
    "negative": "Negative", "NEGATIVE": "Negative", "Negative": "Negative",

    # Neutral variants
    "neytral": "Neutral", "NEYTRAL": "Neutral", "Neytral": "Neutral",
    "neutral": "Neutral", "NEUTRAL": "Neutral", "Neutral": "Neutral",
    "tərəfsiz": "Neutral", "TƏRƏFSİZ": "Neutral",
    "terefsiz": "Neutral", "TEREFSIZ": "Neutral",

    # Numeric labels
    "1": "Positive", "0": "Neutral", "-1": "Negative",
}

def _fold(s: str) -> str:
    s = unicodedata.normalize("NFKD", s.lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.replace("ı", "i").replace("ə", "e")

def normalize_date_text(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip()
    s = re.sub(r"\s+\d{1,2}:\d{2}(:\d{2})?\s*(AM|PM)?$", "", s, flags=re.I)
    s = s.replace("/", "-").replace(".", "-").replace("_", "-").replace("–", "-").replace("—", "-")
    s = re.sub(r"-{2,}", "-", s)
    return s.strip('"').strip("'").strip()

def translate_sentiment(x) -> str:
    if pd.isna(x):
        return "Neutral"
    raw = str(x).strip()
    if not raw:
        return "Neutral"
    key = re.sub(r"\s+", " ", _fold(raw))
    if key in SENTIMENT_MAP:
        return SENTIMENT_MAP[key]
    # Unrecognized → Neutral
    return "Neutral"

def best_col(df: pd.DataFrame, candidates) -> Optional[str]:
    cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        for k, original in cols_lower.items():
            if cand in k:
                return original
    return None

def process_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """Clean a single sheet DataFrame. Returns cleaned DataFrame."""
    df = df.reset_index(drop=True)

    url_col       = best_col(df, ["url", "link", "href", "source"])
    content_col   = best_col(df, ["content", "text", "metn", "mətn", "kontent",
                                   "message", "post", "caption", "description", "body"])
    date_col      = best_col(df, ["date", "tarix", "data", "datetime",
                                   "time", "timestamp", "created", "published",
                                   "posted", "vaxt", "zaman", "created_at",
                                   "publish", "gun", "gün"])
    sentiment_col = best_col(df, ["sentiment", "hiss", "emosiya", "rating",
                                   "tone", "mood", "label", "class"])
    measures_col  = best_col(df, ["measures", "tədbir", "action"])

    # URL və Content tapılmasa → sheet-i skip et
    if not url_col and not content_col:
        raise ValueError(
            f"URL və Content sütunları tapılmadı. "
            f"Mövcud sütunlar: {list(df.columns)}"
        )

    if not url_col:
        url_col = content_col
    if not content_col:
        content_col = url_col

    # Date tapılmasa boş burax
    if date_col:
        normalized = df[date_col].map(normalize_date_text)
        parsed = pd.to_datetime(normalized, errors="coerce", dayfirst=True)
        if parsed.isna().mean() > 0.5:
            alt = pd.to_datetime(normalized, errors="coerce", yearfirst=True)
            if alt.isna().mean() < parsed.isna().mean():
                parsed = alt
    else:
        parsed = pd.Series([pd.NaT] * len(df))

    # Sentiment tapılmasa Neutral yaz
    if sentiment_col:
        sentiments = df[sentiment_col].map(translate_sentiment).values
    else:
        sentiments = ["Neutral"] * len(df)

    out = pd.DataFrame({
        "URL":       df[url_col].fillna("").astype(str).str.strip(),
        "Content":   df[content_col].fillna("").astype(str).str.strip(),
        "Date":      parsed.dt.date.where(parsed.notna(), None),
        "Sentiment": sentiments,
        "_sort":     parsed.values,
    })
    if measures_col:
        out["Measures taken"] = df[measures_col].fillna("").astype(str).str.strip()

    out = (
        out.sort_values("_sort", ascending=True, na_position="last")
           .reset_index(drop=True)
           .drop(columns=["_sort"])
    )
    return out

def process_excel(uploaded_bytes: bytes) -> tuple:
    """Read every sheet, process each separately, write to output with same sheet names."""
    all_sheets: dict = pd.read_excel(
        io.BytesIO(uploaded_bytes), sheet_name=None, dtype=str
    )

    skipped = []
    processed = {}
    passthrough = {}
    for sheet_name, df in all_sheets.items():
        if sheet_name.lower() == "report":
            passthrough[sheet_name] = df
            continue
        try:
            processed[sheet_name] = process_sheet(df)
        except ValueError as e:
            skipped.append(f"⚠️ Sheet '{sheet_name}' skip olundu: {e}")

    if not processed and not passthrough:
        raise ValueError(
            "Heç bir sheet emal oluna bilmədi.\n" + "\n".join(skipped)
        )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl", datetime_format="YYYY-MM-DD") as writer:
        for sheet_name, df in passthrough.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        for sheet_name, cleaned in processed.items():
            cleaned.to_excel(writer, index=False, sheet_name=sheet_name)

            ws = writer.sheets[sheet_name]
            header = {cell.value: cell.column for cell in ws[1]}
            url_col_idx  = header.get("URL")
            date_col_idx = header.get("Date")
            hyperlink_font = Font(color="0563C1", underline="single")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if url_col_idx:
                    cell = row[url_col_idx - 1]
                    url_val = str(cell.value or "").strip()
                    if url_val.startswith("http"):
                        cell.hyperlink = url_val
                        cell.font = hyperlink_font
                if date_col_idx:
                    dcell = row[date_col_idx - 1]
                    dcell.number_format = "YYYY-MM-DD"
    return buf.getvalue(), skipped


st.title("Excel Cleaner")
st.write("Excel yüklə → bütün sheet-lər ayrı-ayrı eşlal olunacaq. Hər sheet-də çıxış: URL, Content, Date (YYYY-MM-DD), Sentiment.")

uploaded = st.file_uploader("Excel faylını seç (.xlsx)", type=["xlsx"])

col1, col2 = st.columns(2)
with col1:
    run_btn = st.button("Təmizlə və hazırla", type="primary")
with col2:
    st.caption("Çıxış: hər sheet ayrıca → URL · Content · Date · Sentiment")

if run_btn:
    if not uploaded:
        st.error("Əvvəl Excel faylını yüklə.")
    else:
        try:
            result, skipped = process_excel(uploaded.getvalue())
            for msg in skipped:
                st.warning(msg)
            st.success("Hazırdır ✅")
            st.download_button(
                label="Cleaned Excel-i yüklə",
                data=result,
                file_name="cleaned_final.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(str(e))